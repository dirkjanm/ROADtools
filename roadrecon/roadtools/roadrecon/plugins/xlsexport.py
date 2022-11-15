"""
Export to Excel file plugin
Contributed by Bastien Cacace (XMCO)

Copyright 2020 - MIT License

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

import argparse
import json
from html import escape
import os
import pprint
import types

from marshmallow import Schema, fields
from marshmallow_sqlalchemy import ModelConverter, SQLAlchemyAutoSchema
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from roadtools.roadlib.metadef.database import (
    User, JSON, Group, DirectoryRole, ServicePrincipal, AppRoleAssignment,
    RoleAssignment, TenantDetail, Application, Device, OAuth2PermissionGrant
)
from roadtools.roadrecon.server import (
    user_schema, device_schema, group_schema, application_schema,
    td_schema, serviceprincipal_schema, users_schema, devices_schema,
    groups_schema, applications_schema, serviceprincipals_schema
)
import roadtools.roadlib.metadef.database as database

# Required property - plugin description
DESCRIPTION = "Export data to an Excel file"

# Plugin properties
SUPPORTED_EXTENSIONS = ['.xls', '.xlsx']


class ExportToFilePlugin():
    """
    Export data to a file (Excel format).
    """

    def __init__(self, session, file, verbose=False):
        self.session = session
        self.file = file
        self.verbose = verbose

    def _print_msg(self, msg):
        if self.verbose:
            print(msg)

    def _create_excel_headers(self, sheet, list_headers):
        i = 1
        for header in list_headers:
            sheet.cell(row=1, column=i).value = header
            i += 1

    def _apply_style_sheet(self, sheet, width):
        # Style
        i = ord('A')
        for header in sheet[1]:
            header.font = Font(bold=True, color="FFFEFB")
            header.fill = PatternFill("solid", fgColor="47402D")
            sheet.column_dimensions[chr(i)].width = width
            i += 1

    def _fill_sheet(self, sheet, assets, fields):
        i = 2
        for asset in assets:
            j = 1
            for field in fields:
                if isinstance(asset, dict):
                    result = asset.get(field)
                else:
                    result = getattr(asset, field)

                if isinstance(result, types.GeneratorType):
                    result = sorted(set(result))
                    result = "\n".join(result)
                if isinstance(result, (list,)) and len(result) == 0:
                    result = ""
                if isinstance(result, (list,)):
                    if isinstance(result[0], (Group,User,ServicePrincipal)):
                        object_name = []
                        for obj in result:
                            object_name.append(obj.displayName)
                        result = "\n".join(object_name)
                    elif isinstance(result[0], (dict,)):
                        result = len(result)
                    else:
                        result = sorted(set(result))
                        result = " ".join(result)
                if isinstance(result, (dict,)):
                    result = json.dumps(result)

                sheet.cell(row=i, column=j).value = result
                j += 1
            i += 1

    def _create_sheet(self, book, name):
        book.create_sheet(name)
        sheet = book[name]
        return sheet

    def get_users(self, book, column_width=40):
        sheet_name = "Users"
        self._print_msg('Export %s info' % sheet_name)

        sheet = self._create_sheet(book, sheet_name)
        self._create_excel_headers(sheet, users_schema.Meta().fields)
        self._apply_style_sheet(sheet, column_width)
        all_users = self.session.query(User).all()
        self._fill_sheet(sheet, all_users, users_schema.Meta().fields)

    def get_devices(self, book, column_width=40):
        sheet_name = "Devices"
        self._print_msg('Export %s info' % sheet_name)

        sheet = self._create_sheet(book, sheet_name)
        self._create_excel_headers(sheet, devices_schema.Meta().fields)
        self._apply_style_sheet(sheet, column_width)
        all_devices = self.session.query(Device).all()
        self._fill_sheet(sheet, all_devices, devices_schema.Meta().fields)

    def get_groups(self, book, column_width=40):
        sheet_name = "Groups"
        self._print_msg('Export %s info' % sheet_name)

        sheet = self._create_sheet(book, sheet_name)
        self._create_excel_headers(sheet, groups_schema.Meta().fields)
        self._apply_style_sheet(sheet, column_width)
        all_groups = self.session.query(Group).all()
        self._fill_sheet(sheet, all_groups, groups_schema.Meta().fields)

    def get_member_of(self, book, column_width=40):
        sheet_name = "MemberOf"
        self._print_msg('Export %s info' % sheet_name)

        members_of = dict()
        hearders = ['objectId', 'displayName', 'memberOf']
        sheet = self._create_sheet(book, "MemberOf")
        self._create_excel_headers(sheet, hearders)
        self._apply_style_sheet(sheet, column_width)
        all_users = self.session.query(User).all()
        self._fill_sheet(sheet, all_users, hearders)

    def get_directory_roles(self, book, column_width=40):
        sheet_name = "Directory roles"
        self._print_msg('Export %s info' % sheet_name)

        fields = (
            'objectId', 'objectType', 'displayName', 'cloudSecurityIdentifier',
            'description', 'isSystem', 'roleDisabled', 'member', 'memberType'
        )
        sheet = self._create_sheet(book, sheet_name)
        self._create_excel_headers(sheet, fields)
        self._apply_style_sheet(sheet, column_width)
        all_directory_roles = self.session.query(DirectoryRole).all()
        directory_roles_by_member = []
        for directory_role in all_directory_roles:
            members = directory_role.memberUsers if directory_role.memberUsers else directory_role.memberServicePrincipals
            for member in members:
                directory_roles_by_member.append({
                    'objectId': directory_role.objectId,
                    'objectType': directory_role.objectType,
                    'displayName': directory_role.displayName,
                    'cloudSecurityIdentifier': directory_role.cloudSecurityIdentifier,
                    'displayName': directory_role.displayName,
                    'description': directory_role.description,
                    'member': member.displayName,
                    'memberType': member.objectType,
                    'isSystem': directory_role.isSystem,
                    'roleDisabled': directory_role.roleDisabled
                })

        self._fill_sheet(sheet, directory_roles_by_member, fields)

    def get_applications(self, book, column_width=40):
        sheet_name = "Applications"
        self._print_msg('Export %s info' % sheet_name)

        fields = (
            'objectId', 'objectType', 'displayName', 'appId',
            'oauth2AllowIdTokenImplicitFlow', 'availableToOtherTenants',
            'publisherDomain', 'replyUrls', 'appRoles', 'publicClient',
            'oauth2AllowImplicitFlow', 'oauth2Permissions', 'homepage',
            'passwordCredentials', 'keyCredentials', 'ownerUsers',
            'ownerServicePrincipals'
        )
        sheet = self._create_sheet(book, sheet_name)
        self._create_excel_headers(sheet, fields)
        self._apply_style_sheet(sheet, column_width)
        all_applications = self.session.query(Application).all()
        self._fill_sheet(sheet, all_applications, fields)

    def get_service_principals(self, book, column_width=40):
        sheet_name = "Service principals"
        self._print_msg('Export %s info' % sheet_name)

        fields = (
            'objectId', 'objectType', 'displayName', 'appDisplayName',
            'appId', 'publisherName', 'replyUrls', 'appRoles',
            'microsoftFirstParty', 'oauth2Permissions', 'passwordCredentials',
            'keyCredentials', 'ownerUsers', 'ownerServicePrincipals',
            'accountEnabled', 'servicePrincipalType'
        )
        sheet = self._create_sheet(book, sheet_name)
        self._create_excel_headers(sheet, fields)
        self._apply_style_sheet(sheet, column_width)
        all_service_principal = self.session.query(ServicePrincipal).all()
        self._fill_sheet(sheet, all_service_principal, fields)

    def get_app_roles(self, book, column_width=40):
        sheet_name = "Applications roles"
        self._print_msg('Export %s info' % sheet_name)

        sheet = self._create_sheet(book, sheet_name)
        fields = ('objid', 'ptype', 'pname', 'app', 'value', 'desc', 'spid')
        self._create_excel_headers(sheet, fields)
        self._apply_style_sheet(sheet, column_width)
        approles = []
        for ar in self.session.query(AppRoleAssignment).all():
            rsp = self.session.query(ServicePrincipal).get(ar.resourceId)
            if ar.principalType == 'ServicePrincipal':
                sp = self.session.query(ServicePrincipal).get(ar.principalId)
            if ar.principalType == 'User':
                sp = self.session.query(User).get(ar.principalId)
            if ar.principalType == 'Group':
                sp = self.session.query(Group).get(ar.principalId)
            if not sp:
                self._print_msg('Could not resolve service principal for approle {0}'.format(str(ar)))
                continue
            if ar.id == '00000000-0000-0000-0000-000000000000':
                approles.append({
                    'objid': sp.objectId,
                    'ptype': ar.principalType,
                    'pname': sp.displayName,
                    'app': ar.resourceDisplayName,
                    'value': 'Default',
                    'desc': 'Default Role',
                    'spid': ar.resourceId,
                })
            else:
                for approle in rsp.appRoles:
                    if approle['id'] == ar.id:
                        approles.append({
                            'objid': sp.objectId,
                            'ptype': ar.principalType,
                            'pname': sp.displayName,
                            'app': ar.resourceDisplayName,
                            'value': approle['value'],
                            'desc': approle['displayName'],
                            'spid': ar.resourceId,
                        })
        self._fill_sheet(sheet, approles, fields)

    def get_oauth2_permissions(self, book, column_width=40):
        sheet_name = "Oauth2 permissions"
        self._print_msg('Export %s info' % sheet_name)

        sheet = self._create_sheet(book, sheet_name)
        oauth2permissions = []
        fields = (
            'type', 'userid', 'userdisplayname', 'targetapplication', 'targetspobjectid',
            'sourceapplication', 'sourcespobjectid', 'expiry', 'scope'
        )
        self._create_excel_headers(sheet, fields)
        self._apply_style_sheet(sheet, column_width)
        for permgrant in self.session.query(OAuth2PermissionGrant).all():
            grant = {}
            rsp = self.session.query(ServicePrincipal).get(permgrant.clientId)
            if permgrant.consentType == 'Principal':
                grant['type'] = 'user'
                user = self.session.query(User).get(permgrant.principalId)
                grant['userid'] = user.objectId
                grant['userdisplayname'] = user.displayName
            else:
                grant['type'] = 'all'
                grant['userid'] = None
                grant['userdisplayname'] = None

            targetapp = self.session.query(ServicePrincipal).get(permgrant.resourceId)
            grant['targetapplication'] = targetapp.displayName
            grant['targetspobjectid'] = targetapp.objectId
            grant['sourceapplication'] = rsp.displayName
            grant['sourcespobjectid'] = rsp.objectId
            grant['expiry'] = permgrant.expiryTime
            grant['scope'] = permgrant.scope
            oauth2permissions.append(grant)

        self._fill_sheet(sheet, oauth2permissions, fields)

    def get_mfa(self, book, column_width=40):
        sheet_name = "MFA"
        self._print_msg('Export %s info' % sheet_name)

        sheet = self._create_sheet(book, sheet_name)
        fields = (
            'objectId', 'displayName', 'mfamethods', 'accountEnabled', 'has_app',
            'has_phonenr', 'has_fido', 'encryptedPinHash', 'encryptedPinHashHistory',
            'methods', 'oathTokenMetadata', 'requirements', 'phoneAppDetails',
            'proofupTime', 'verificationDetail', 'requirements'
        )
        self._create_excel_headers(sheet, fields)
        self._apply_style_sheet(sheet, column_width)
        all_mfa = self.session.query(User).all()
        mfa = []
        for user in all_mfa:
            mfa_methods = len(user.strongAuthenticationDetail['methods'])
            methods = [method['methodType'] for method in user.strongAuthenticationDetail['methods']]
            has_app = 'PhoneAppOTP' in methods or 'PhoneAppNotification' in methods
            has_phonenr = 'OneWaySms' in methods or 'TwoWayVoiceMobile' in methods
            has_fido = 'FIDO' in [key['usage'] for key in user.searchableDeviceKey]
            mfa.append({
                'objectId': user.objectId,
                'displayName': user.displayName,
                'mfamethods': mfa_methods,
                'accountEnabled': user.accountEnabled,
                'has_app': has_app,
                'has_phonenr': has_phonenr,
                'has_fido': has_fido,
                'encryptedPinHash': user.strongAuthenticationDetail['encryptedPinHash'],
                'encryptedPinHashHistory': user.strongAuthenticationDetail['encryptedPinHashHistory'],
                'methods': methods,
                'oathTokenMetadata': user.strongAuthenticationDetail['oathTokenMetadata'],
                'requirements': user.strongAuthenticationDetail['requirements'],
                'phoneAppDetails': user.strongAuthenticationDetail['phoneAppDetails'],
                'proofupTime': user.strongAuthenticationDetail['proofupTime'],
                'verificationDetail': user.strongAuthenticationDetail['verificationDetail'],
                'requirements': user.strongAuthenticationDetail['requirements']
            })

        self._fill_sheet(sheet, mfa, fields)

    def main(self):
        wb = Workbook()
        wb.remove(wb.active)

        self.get_users(wb)
        self.get_devices(wb)
        self.get_groups(wb)
        self.get_member_of(wb)
        self.get_directory_roles(wb)
        self.get_applications(wb)
        self.get_service_principals(wb)
        self.get_app_roles(wb)
        self.get_oauth2_permissions(wb)
        self.get_mfa(wb)
        wb.save(self.file)


def create_args_parser():
    parser = argparse.ArgumentParser(
        add_help=True,
        description=DESCRIPTION,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        '-d', '--database',
        action='store',
        help='Database file. Can be the local database name for SQLite, or an SQLAlchemy '
             'compatible URL such as postgresql+psycopg2://dirkjan@/roadtools',
        default='roadrecon.db'
    )
    add_args(parser)
    return parser


def add_args(parser):
    parser.add_argument(
        '-f', '--file',
        action='store',
        help='Output excel file (default: data.xlsx)',
        default='data.xls'
    )
    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Also print details to the console',
        required=False
    )


def main(args=None):
    if args is None:
        parser = create_args_parser()
        args = parser.parse_args()

    db_url = database.parse_db_argument(args.database)
    session = database.get_session(database.init(dburl=db_url))
    filename, file_extension = os.path.splitext(args.file)
    if not file_extension:
        file_extension = '.xlsx'
    if file_extension not in SUPPORTED_EXTENSIONS:
        print("%s is not a supported extention. Only %s are supported" % (file_extension, ', '.join(SUPPORTED_EXTENSIONS)))
        return

    plugin = ExportToFilePlugin(session, filename + file_extension, verbose=args.verbose)
    plugin.main()
    print("Data have been exported to the %s%s file" % (filename, file_extension))

if __name__ == '__main__':
    main()
